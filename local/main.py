import cv2
import pickle
import face_recognition
import numpy as np
import time
import math
import cvzone
from ultralytics import YOLO
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from datetime import datetime
import os

# --- Setup Excel ---
excel_file = 'DataPekerja.xlsx'

def setup_excel():
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime("%Y-%m-%d")
        ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi'])
        wb.save(excel_file)
        wb.close()
    else:
        wb = openpyxl.load_workbook(excel_file)
        sheet_name = datetime.now().strftime("%Y-%m-%d")
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi'])
        wb.save(excel_file)
        wb.close()

setup_excel()

# --- Load Model dan Encoding Wajah ---
modelhelm = YOLO("helm.pt")
modelrompi = YOLO("rompi.pt")
modelsepatu = YOLO("sepatu.pt")
modelsarungtangan = YOLO("sarungtangan.pt")

with open('EncodeFile.p', 'rb') as file:
    encodeListKnownWithIds = pickle.load(file)
encodeListKnown, knownNames = encodeListKnownWithIds

# --- Fungsi Countdown dengan Progress Bar ---
def show_countdown_gui(message, seconds=3):
    root = tk.Tk()
    root.title("Countdown Persiapan")
    root.geometry("400x200")

    label = tk.Label(root, text="", font=("Helvetica", 20))
    label.pack(pady=20)

    progress = ttk.Progressbar(root, length=300, mode='determinate')
    progress.pack(pady=20)
    progress['maximum'] = seconds

    def update_countdown(count):
        if count >= 0:
            label.config(text=f"{message}: {count} detik")
            progress['value'] = seconds - count
            root.update()
            root.after(1000, update_countdown, count-1)
        else:
            root.destroy()

    update_countdown(seconds)
    root.mainloop()

# --- Fungsi Deteksi Wajah ---
def detect_face():
    print("\nPersiapan Deteksi Wajah...")
    show_countdown_gui("Bersiap untuk deteksi wajah", seconds=3)


    cap = cv2.VideoCapture(0)
    cap.set(3, 1280)
    cap.set(4, 720)

    detected_name = None

    while True:
        success, img = cap.read()
        if not success:
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        faceCurFrame = face_recognition.face_locations(imgS)
        encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

        if encodeCurFrame:
            encodeFace = encodeCurFrame[0]
            faceLoc = faceCurFrame[0]

            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis) if faceDis.size > 0 else -1

            if matchIndex != -1 and matches[matchIndex]:
                detected_name = knownNames[matchIndex]
                print(f'Detected: {detected_name}')
                cap.release()
                cv2.destroyAllWindows()
                return detected_name

        cv2.imshow("Face Recognition", img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    return None

# --- Fungsi Deteksi APD ---
def detect_apd():
    cap = cv2.VideoCapture(0)
    cap.set(3, 1280)
    cap.set(4, 720)

    detected_items = set()

    print("\nPersiapan Deteksi APD...")
    show_countdown_gui("Bersiap untuk deteksi APD", seconds=3)

    start_time = time.time()
    timeout = 10  # batas deteksi dalam detik

    while True:
        success, img = cap.read()
        if not success:
            break

        # Proses masing-masing model YOLO
        for model, label in [
            (modelhelm, "Helm"),
            (modelrompi, "Rompi"),
            (modelsepatu, "Sepatu"),
            (modelsarungtangan, "Sarung Tangan")
        ]:
            results = model(img, stream=True)
            for r in results:
                boxes = r.boxes
                for box in boxes:
                    conf = box.conf[0]
                    if conf > 0.65:
                        detected_items.add(label)
                        x1, y1, x2, y2 = box.xyxy[0]
                        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                        w, h = x2 - x1, y2 - y1
                        cvzone.cornerRect(img, (x1, y1, w, h))
                        conf_text = f'{label} {math.ceil((conf * 100)) / 100}'
                        cvzone.putTextRect(img, conf_text, (max(0, x1), max(35, y1)), scale=1, thickness=1)

        cv2.imshow("APD Detection", img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

        if len(detected_items) >= 4 or (time.time() - start_time > timeout):
            time.sleep(2)
            cap.release()
            cv2.destroyAllWindows()
            return detected_items

    cap.release()
    cv2.destroyAllWindows()
    return detected_items

# --- Fungsi Simpan ke Excel ---
def save_to_excel(name, status, apd_list):
    wb = openpyxl.load_workbook(excel_file)
    sheet_name = datetime.now().strftime("%Y-%m-%d")
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi'])
    ws = wb[sheet_name]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    apd_text = ', '.join(apd_list)
    ws.append([name, status, now, apd_text])

    wb.save(excel_file)
    wb.close()

# --- Fungsi Konfirmasi Deteksi ---
def confirm_detection(name):
    detected_items = detect_apd()

    status = "Diizinkan" if len(detected_items) == 4 else "Tidak Diizinkan"

    # Buat Window konfirmasi
    root = tk.Tk()
    root.title("Konfirmasi Deteksi")
    root.geometry("400x250")

    label = tk.Label(root, text=f"Nama: {name}\nAPD Terdeteksi: {', '.join(detected_items)}\nStatus Kerja: {status}\n\nApakah deteksi sudah benar?", font=("Helvetica", 10))
    label.pack(pady=10)

    def sudah_benar():
        save_to_excel(name, status, detected_items)
        messagebox.showinfo("Sukses", "Data berhasil disimpan!")
        root.destroy()

    def ulangi_apd():
        root.destroy()
        print("Mengulang hanya deteksi APD...")
        confirm_detection(name)

    def ulangi_wajah():
        root.destroy()
        print("Kembali ke deteksi wajah...")
        run_process()

    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)

    btn_sudah = tk.Button(button_frame, text="Sudah Benar", command=sudah_benar, width=15, bg="green", fg="white")
    btn_sudah.grid(row=0, column=0, padx=5)

    btn_ulangi_apd = tk.Button(button_frame, text="Ulangi Deteksi APD", command=ulangi_apd, width=18, bg="orange")
    btn_ulangi_apd.grid(row=0, column=1, padx=5)

    btn_ulangi_wajah = tk.Button(button_frame, text="Kembali Deteksi Wajah", command=ulangi_wajah, width=20, bg="red", fg="white")
    btn_ulangi_wajah.grid(row=1, column=0, columnspan=2, pady=10)

    root.mainloop()

# --- Fungsi Main Loop ---
def run_process():
    name = detect_face()
    if name:
        confirm_detection(name)
    else:
        print("Wajah tidak terdeteksi!")

# --- Jalankan Program ---
if __name__ == "__main__":
    run_process()
