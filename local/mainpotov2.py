import cv2
import numpy as np
import math
import cvzone
import pickle
import face_recognition
from ultralytics import YOLO
import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from datetime import datetime
import os
import shutil
import re

# --- Setup Folder Laporan ---
def setup_folder():
    today = datetime.now().strftime("%Y-%m-%d")
    folder_path = f"Laporan/{today}"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

# --- Setup Excel ---
def setup_excel(folder_path):
    excel_file = os.path.join(folder_path, 'DataPekerja.xlsx')
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime("%Y-%m-%d")
        ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi', 'Foto APD'])
        wb.save(excel_file)
        wb.close()
    return excel_file

# --- Load Face Encoding ---
with open('EncodeFile.p', 'rb') as file:
    encodeListKnownWithIds = pickle.load(file)
encodeListKnown, knownNames = encodeListKnownWithIds

# --- Load YOLO Model ---
modelAPD = YOLO("kayaknyaoke.pt")
classNames = ["Helm", "Safety Boots", "Rompi", "Sarung Tangan"]

# --- Fungsi Deteksi Wajah dari Gambar ---
def detect_face_from_image(img):
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    faceCurFrame = face_recognition.face_locations(imgS)
    encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

    if encodeCurFrame:
        encodeFace = encodeCurFrame[0]
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis) if faceDis.size > 0 else -1

        if matchIndex != -1 and matches[matchIndex]:
            detected_name = knownNames[matchIndex]
            print(f'Detected: {detected_name}')
            return detected_name

    print("Wajah tidak dikenali.")
    return None

# --- Fungsi Deteksi APD dari Gambar ---
def detect_apd_from_image(img):
    detected_items = set()
    results = modelAPD(img, stream=True)

    for r in results:
        boxes = r.boxes
        for box in boxes:
            conf = box.conf[0]
            if conf > 0.65:
                cls = int(box.cls[0])
                detected_items.add(classNames[cls])

                x1, y1, x2, y2 = box.xyxy[0]
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                w, h = x2 - x1, y2 - y1
                cvzone.cornerRect(img, (x1, y1, w, h))
                conf = math.ceil((conf * 100)) / 100
                cvzone.putTextRect(img, f'{classNames[cls]} {conf}', (max(0, x1), max(35, y1)), scale=1, thickness=1)

    cv2.imshow("APD Detection", img)
    cv2.waitKey(3000)
    cv2.destroyAllWindows()

    return detected_items

# --- Fungsi Simpan ke Excel ---
def save_to_excel(name, status, apd_list, photo_filename, excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet_name = datetime.now().strftime("%Y-%m-%d")
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi', 'Foto APD'])
    ws = wb[sheet_name]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    apd_text = ', '.join(apd_list)
    ws.append([name, status, now, apd_text, photo_filename])

    wb.save(excel_file)
    wb.close()

# --- Fungsi Membersihkan Nama File ---
def clean_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "", name)

# --- Fungsi Konfirmasi Deteksi ---
def confirm_detection(name, detected_items, img_path, folder_path, excel_file):
    status = "Diizinkan" if len(detected_items) == 4 else "Tidak Diizinkan"

    ext = os.path.splitext(img_path)[1]
    safe_name = clean_filename(name)
    photo_filename = f"{safe_name}_APD{ext}"
    target_photo_path = os.path.join(folder_path, photo_filename)
    shutil.copy(img_path, target_photo_path)

    root = tk.Tk()
    root.title("Konfirmasi Deteksi")
    root.geometry("400x250")

    label = tk.Label(root, text=f"Nama: {name}\nAPD Terdeteksi: {', '.join(detected_items)}\nStatus Kerja: {status}\n\nApakah deteksi sudah benar?", font=("Helvetica", 10))
    label.pack(pady=10)

    def sudah_benar():
        save_to_excel(name, status, detected_items, photo_filename, excel_file)
        messagebox.showinfo("Sukses", "Data berhasil disimpan!")
        root.destroy()

    def ulangi():
        root.destroy()
        print("Kembali ke upload foto...")
        run_process()

    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)

    btn_sudah = tk.Button(button_frame, text="Sudah Benar", command=sudah_benar, width=15, bg="green", fg="white")
    btn_sudah.grid(row=0, column=0, padx=5)

    btn_ulangi = tk.Button(button_frame, text="Ulangi", command=ulangi, width=18, bg="red", fg="white")
    btn_ulangi.grid(row=0, column=1, padx=5)

    root.mainloop()

# --- Main Loop ---
def run_process():
    folder_path = setup_folder()
    excel_file = setup_excel(folder_path)

    img_path = filedialog.askopenfilename(title="Pilih Foto untuk Deteksi Wajah dan APD")
    if not img_path:
        print("Tidak ada foto dipilih!")
        return

    img = cv2.imread(img_path)
    if img is None:
        print("Gagal membaca gambar!")
        return

    name = detect_face_from_image(img)
    if name:
        detected_items = detect_apd_from_image(img)
        confirm_detection(name, detected_items, img_path, folder_path, excel_file)
    else:
        messagebox.showerror("Error", "Wajah tidak dikenali, silakan coba lagi!")

# --- Jalankan Program ---
if __name__ == "__main__":
    run_process()
