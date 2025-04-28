import cv2
import numpy as np
import time
import math
import cvzone
from ultralytics import YOLO
import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
import os
import shutil

# --- Setup Excel Folder ---
def setup_folder():
    today = datetime.now().strftime("%Y-%m-%d")
    folder_path = f"Laporan/{today}"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

# --- Fungsi Membuat File Excel ---
def setup_excel(folder_path):
    excel_file = os.path.join(folder_path, 'DataPekerja.xlsx')
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = datetime.now().strftime("%Y-%m-%d")
        ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi', 'Foto APD'])
        wb.save(excel_file)
        wb.close()
    return excel_file

# --- Load Model APD ---
modelAPD = YOLO("tes2.pt")
classNames = ["Helm", "Sarung Tangan", "Safety Boots", "Rompi"]

# --- Fungsi Deteksi APD dari Foto ---
def detect_apd_from_image(image_path):
    img = cv2.imread(image_path)
    if img is None:
        print("Gagal membaca gambar APD!")
        return set()

    detected_items = set()
    results = modelAPD(img, stream=True)

    for r in results:
        boxes = r.boxes
        for box in boxes:
            conf = box.conf[0]
            if conf > 0.78:
                cls = int(box.cls[0])
                detected_items.add(classNames[cls])

                x1, y1, x2, y2 = box.xyxy[0]
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                w, h = x2 - x1, y2 - y1
                cvzone.cornerRect(img, (x1, y1, w, h))
                conf = math.ceil((conf * 100)) / 100
                cvzone.putTextRect(img, f'{classNames[cls]} {conf}', (max(0, x1), max(35, y1)), scale=1, thickness=1)

    cv2.imshow("APD Detection", img)
    cv2.waitKey(3000)  # tampilkan hasil 3 detik
    cv2.destroyAllWindows()

    return detected_items

# --- Fungsi Simpan ke Excel + Insert Gambar ---
def save_to_excel(name, status, apd_list, photo_path, excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet_name = datetime.now().strftime("%Y-%m-%d")
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Nama', 'Status Kerja', 'Waktu Masuk', 'APD Terdeteksi', 'Foto APD'])
    ws = wb[sheet_name]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    apd_text = ', '.join(apd_list)
    ws.append([name, status, now, apd_text])

    # Menyisipkan gambar ke kolom E (Foto APD)
    img = ExcelImage(photo_path)
    img.width = 100  # Atur ukuran agar tidak terlalu besar
    img.height = 100

    img_cell = f'E{ws.max_row}'
    ws.add_image(img, img_cell)

    wb.save(excel_file)
    wb.close()

# --- Fungsi Konfirmasi Deteksi ---
def confirm_detection(name, folder_path, excel_file):
    apd_image_path = filedialog.askopenfilename(title="Pilih Foto untuk Deteksi APD")
    if not apd_image_path:
        print("Tidak ada foto APD dipilih!")
        return

    detected_items = detect_apd_from_image(apd_image_path)

    status = "Diizinkan" if len(detected_items) == 4 else "Tidak Diizinkan"

    # Copy foto ke folder laporan
    photo_filename = os.path.basename(apd_image_path)
    target_photo_path = os.path.join(folder_path, photo_filename)
    shutil.copy(apd_image_path, target_photo_path)

    root = tk.Tk()
    root.title("Konfirmasi Deteksi")
    root.geometry("400x250")

    label = tk.Label(root, text=f"Nama: {name}\nAPD Terdeteksi: {', '.join(detected_items)}\nStatus Kerja: {status}\n\nApakah deteksi sudah benar?", font=("Helvetica", 10))
    label.pack(pady=10)

    def sudah_benar():
        save_to_excel(name, status, detected_items, target_photo_path, excel_file)
        messagebox.showinfo("Sukses", "Data berhasil disimpan!")
        root.destroy()

    def ulangi_apd():
        root.destroy()
        print("Mengulang hanya deteksi APD...")
        confirm_detection(name, folder_path, excel_file)

    def ulangi_nama():
        root.destroy()
        print("Kembali ke input nama...")
        run_process()

    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)

    btn_sudah = tk.Button(button_frame, text="Sudah Benar", command=sudah_benar, width=15, bg="green", fg="white")
    btn_sudah.grid(row=0, column=0, padx=5)

    btn_ulangi_apd = tk.Button(button_frame, text="Ulangi Deteksi APD", command=ulangi_apd, width=18, bg="orange")
    btn_ulangi_apd.grid(row=0, column=1, padx=5)

    btn_ulangi_nama = tk.Button(button_frame, text="Kembali Input Nama", command=ulangi_nama, width=20, bg="red", fg="white")
    btn_ulangi_nama.grid(row=1, column=0, columnspan=2, pady=10)

    root.mainloop()

# --- Fungsi Input Nama dan Main Loop ---
def run_process():
    folder_path = setup_folder()
    excel_file = setup_excel(folder_path)

    def submit_name():
        entered_name = entry_name.get().strip()
        if entered_name:
            root.destroy()
            confirm_detection(entered_name, folder_path, excel_file)
        else:
            messagebox.showerror("Error", "Nama tidak boleh kosong!")

    root = tk.Tk()
    root.title("Input Nama Pekerja")
    root.geometry("400x200")

    label = tk.Label(root, text="Masukkan Nama:", font=("Helvetica", 14))
    label.pack(pady=20)

    entry_name = tk.Entry(root, font=("Helvetica", 14))
    entry_name.pack(pady=10)

    btn_submit = tk.Button(root, text="Lanjut", command=submit_name, font=("Helvetica", 12), bg="blue", fg="white")
    btn_submit.pack(pady=10)

    root.mainloop()

# --- Jalankan Program ---
if __name__ == "__main__":
    run_process()
